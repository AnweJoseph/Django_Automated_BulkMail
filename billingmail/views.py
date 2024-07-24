from django.shortcuts import render, redirect
from django.core.mail import EmailMessage, send_mail
import openpyxl, smtplib, time
from django.conf import settings


# Create your views here.

def send_mail(request):
    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook('billsRecord.xlsx')
    sheet = wb['Sheet1']
    lastCol = sheet.max_column

    unpaidMembers = {}
    for r in range(2, sheet.max_row):
        payment = sheet.cell(row=r, column=lastCol).value
        if payment != 'paid':
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            unpaidMembers[name] = email

    from_email = settings.DEFAULT_FROM_EMAIL
    if request.method == 'POST':
        
        for client_mail in unpaidMembers.items():
            if client_mail:
                
                recipient = list(client_mail)
                subject = request.POST['subject']
                message = request.POST['message']
                print(recipient)

                mail = EmailMessage(subject,message,from_email,recipient)
                mail.send(fail_silently=True)
                
    return render(request, 'send_mail.html')
